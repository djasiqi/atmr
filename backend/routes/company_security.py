"""Routes API pour l'onglet Securite V2 : audit logs, export, policy."""

from __future__ import annotations

import hashlib
import io
import logging
from contextlib import suppress
from datetime import UTC, datetime, timedelta

from flask import Response, request
from flask_jwt_extended import jwt_required
from flask_restx import Namespace, Resource, fields

from ext import db, limiter, role_required
from models.enums import UserRole
from routes.companies import get_company_from_token
from shared.error_handlers import APIErrorHandler

logger = logging.getLogger(__name__)

security_ns = Namespace(
    "company-security",
    description="Securite entreprise : audit logs, export, politique",
)

# ==================== Labels FR ====================

ACTION_LABELS_FR: dict[str, str] = {
    "login_success": "Connexion",
    "user_login": "Connexion",
    "login_failed": "Tentative de connexion échouée",
    "user_login_failed": "Tentative de connexion échouée",
    "user_logout": "Déconnexion",
    "session_revoked": "Session révoquée",
    "sessions_bulk_revoked": "Sessions révoquées en masse",
    "settings_updated": "Paramètres modifiés",
    "billing_settings_updated": "Paramètres de facturation modifiés",
    "booking_cancelled": "Réservation annulée",
    "booking_created": "Réservation créée",
    "booking_modified": "Réservation modifiée",
    "booking_created_from_request": "Réservation créée (depuis demande)",
    "dispatch_complete": "Dispatch terminé",
    "audit_log_exported": "Journal exporté",
    "client_created": "Client ajouté",
    "client_modified": "Client modifié",
    "driver_assigned": "Chauffeur assigné",
    "invoice_generated": "Facture générée",
    "totp_enabled": "2FA activée",
    "totp_disabled": "2FA désactivée",
    "totp_challenge_failed": "Tentative 2FA échouée",
    "security_policy_updated": "Politique de sécurité modifiée",
    "password_changed": "Mot de passe modifié",
    "data_access": "Accès aux données",
    "user_invited": "Utilisateur invité",
    "recovery_codes_regenerated": "Codes de secours régénérés",
    "request_converted": "Demande convertie en réservation",
    "offer_accepted": "Offre acceptée",
    "offer_rejected": "Offre refusée",
    "transport_request_created": "Demande de transport créée",
    "transport_request_cancelled": "Demande de transport annulée",
    "patient_created": "Patient ajouté",
    "patient_updated": "Patient modifié",
}

CATEGORY_LABELS_FR: dict[str, str] = {
    "security": "Sécurité",
    "dispatch": "Dispatch",
    "billing": "Facturation",
    "data": "Données",
    "operations": "Opérations",
    "settings": "Paramètres",
    "institution": "Institution",
}

RESULT_LABELS_FR: dict[str, str] = {
    "success": "Succès",
    "failure": "Échec",
    "partial": "Partiel",
}

# ==================== RESTX Models ====================

audit_log_item = security_ns.model(
    "AuditLogItem",
    {
        "id": fields.Integer,
        "created_at": fields.DateTime,
        "action_type": fields.String(description="Type brut, traduit cote frontend (G3)"),
        "action_category": fields.String,
        "result_status": fields.String,
        "resource_type": fields.String,
        "resource_id": fields.String,
        "actor_username": fields.String,
        "actor_user_type": fields.String,
        "ip_masked": fields.String(description="IP masquee, jamais complete"),
        "device": fields.String(description="Parse depuis user_agent"),
    },
)

audit_logs_response = security_ns.model(
    "AuditLogsResponse",
    {
        "logs": fields.List(fields.Nested(audit_log_item)),
        "total": fields.Integer,
        "page": fields.Integer,
        "per_page": fields.Integer,
        "has_more": fields.Boolean,
    },
)


def _serialize_audit_log(log) -> dict[str, object]:
    """Serialize un AuditLog avec IP masquee et device parse."""
    from models import User
    from shared.security_helpers import mask_ip, parse_device

    actor_username = None
    actor_user_type = log.user_type
    if log.user_id:
        user = User.query.get(log.user_id)
        if user:
            actor_username = user.username or user.email

    return {
        "id": log.id,
        "created_at": log.created_at.isoformat() if log.created_at else None,
        "action_type": log.action_type,
        "action_category": log.action_category,
        "result_status": log.result_status,
        "resource_type": getattr(log, "resource_type", None),
        "resource_id": getattr(log, "resource_id", None),
        "actor_username": actor_username,
        "actor_user_type": actor_user_type,
        "ip_masked": mask_ip(log.ip_address),
        "device": parse_device(log.user_agent),
    }


# ==================== Audit Logs ====================

@security_ns.route("/audit-logs")
class AuditLogList(Resource):
    @jwt_required()
    @role_required(UserRole.company)
    @security_ns.marshal_with(audit_logs_response)
    def get(self):
        """Liste paginee des audit logs de l'entreprise."""
        company, err, _code = get_company_from_token()
        if err:
            return APIErrorHandler.handle_not_found("Company", None, logger)
        if not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        if request.args.get("company_id"):
            logger.warning(
                "[Security] company_id param ignored in audit-logs request from company %s",
                company.id,
            )

        from security.audit_log import AuditLog

        page = min(max(int(request.args.get("page", 1)), 1), 1000)
        per_page = min(max(int(request.args.get("per_page", 20)), 1), 50)
        category = request.args.get("category")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")

        query = AuditLog.query.filter(AuditLog.company_id == company.id)

        if category and category != "all":
            query = query.filter(AuditLog.action_category == category)
        if date_from:
            try:
                dt_from = datetime.fromisoformat(date_from.replace("Z", "+00:00"))
                query = query.filter(AuditLog.created_at >= dt_from)
            except (ValueError, TypeError):
                pass
        if date_to:
            try:
                dt_to = datetime.fromisoformat(date_to.replace("Z", "+00:00"))
                query = query.filter(AuditLog.created_at <= dt_to)
            except (ValueError, TypeError):
                pass

        total = query.count()
        logs = (
            query.order_by(AuditLog.created_at.desc())
            .offset((page - 1) * per_page)
            .limit(per_page)
            .all()
        )

        return {
            "logs": [_serialize_audit_log(log) for log in logs],
            "total": total,
            "page": page,
            "per_page": per_page,
            "has_more": (page * per_page) < total,
        }


# ==================== Export ====================

_BOM_UTF8 = "\ufeff"
_CSV_HEADERS = [
    "Date", "Heure", "Utilisateur", "Action", "Catégorie",
    "Résultat", "Adresse IP", "Appareil",
]
_CSV_HEADER_LINE = ";".join(_CSV_HEADERS) + "\n"


def _resolve_actor(log) -> str:
    from models import User

    if log.user_id:
        user = User.query.get(log.user_id)
        if user:
            return user.username or user.email or ""
    return ""


def _format_csv_row(log) -> str:
    from shared.security_helpers import mask_ip, parse_device

    actor = _resolve_actor(log)
    dt = log.created_at
    date_str = dt.strftime("%d.%m.%Y") if dt else ""
    time_str = dt.strftime("%H:%M") if dt else ""
    action_label = ACTION_LABELS_FR.get(log.action_type, log.action_type)
    cat_label = CATEGORY_LABELS_FR.get(log.action_category, log.action_category or "")
    result_label = RESULT_LABELS_FR.get(log.result_status, log.result_status or "")

    return (
        f"{date_str};{time_str};{actor};{action_label};"
        f"{cat_label};{result_label};{mask_ip(log.ip_address)};{parse_device(log.user_agent)}\n"
    )


def _build_export_query(company, category, date_from, date_to, limit):
    """Construit la query filtrée pour l'export."""
    from security.audit_log import AuditLog

    query = AuditLog.query.filter(AuditLog.company_id == company.id)

    if category and category != "all":
        query = query.filter(AuditLog.action_category == category)

    now = datetime.now(UTC)
    is_institution = False
    max_days = 365 if is_institution else 90
    max_rows = 50000 if is_institution else 10000
    limit = min(limit, max_rows)
    default_from = now - timedelta(days=max_days)

    if date_from:
        try:
            dt_from = datetime.fromisoformat(date_from.replace("Z", "+00:00"))
            if (now - dt_from).days > max_days:
                dt_from = default_from
            query = query.filter(AuditLog.created_at >= dt_from)
        except (ValueError, TypeError):
            query = query.filter(AuditLog.created_at >= default_from)
    else:
        query = query.filter(AuditLog.created_at >= default_from)

    if date_to:
        try:
            dt_to = datetime.fromisoformat(date_to.replace("Z", "+00:00"))
            query = query.filter(AuditLog.created_at <= dt_to)
        except (ValueError, TypeError):
            pass

    return query.order_by(AuditLog.created_at.desc()).limit(limit), now, is_institution


_XL_HEADERS = [
    "Date", "Heure", "Utilisateur", "Action", "Catégorie",
    "Résultat", "Adresse IP", "Appareil",
]
_XL_COL_WIDTHS = [14, 10, 24, 40, 16, 12, 20, 26]

_COLOR_TEAL = "00796B"
_COLOR_SLATE = "64748B"
_COLOR_DARK_SLATE = "334155"
_COLOR_ROW_ALT = "F0FDFA"
_COLOR_GREEN = "059669"
_COLOR_RED = "C62828"
_COLOR_BLUE = "2563EB"


def _generate_excel(query, company_name: str, now: datetime, total_count: int) -> bytes:
    """Genere un fichier Excel (.xlsx) aligne sur le design institution."""
    from openpyxl import Workbook  # noqa: I001
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from shared.security_helpers import mask_ip, parse_device

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Journal d'activité"
    last_col = len(_XL_HEADERS)
    last_col_letter = get_column_letter(last_col)

    title_font = Font(name="Calibri", bold=True, size=16, color=_COLOR_TEAL)
    subtitle_font = Font(name="Calibri", size=10, color=_COLOR_SLATE)
    summary_font = Font(name="Calibri", bold=True, size=10, color=_COLOR_DARK_SLATE)
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color=_COLOR_TEAL, end_color=_COLOR_TEAL, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center")
    data_font = Font(name="Calibri", size=10.5)
    data_font_secondary = Font(name="Calibri", size=10.5, color=_COLOR_SLATE)
    data_font_action = Font(name="Calibri", bold=True, size=10.5)
    result_font_success = Font(name="Calibri", bold=True, size=10.5, color=_COLOR_GREEN)
    result_font_failure = Font(name="Calibri", bold=True, size=10.5, color=_COLOR_RED)
    result_font_other = Font(name="Calibri", bold=True, size=10.5, color=_COLOR_BLUE)
    even_fill = PatternFill(start_color=_COLOR_ROW_ALT, end_color=_COLOR_ROW_ALT, fill_type="solid")

    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = f"Journal d'activité — {company_name}"
    c.font = title_font
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{last_col_letter}2")
    c = ws["A2"]
    c.value = f"Généré le {now.strftime('%d.%m.%Y')} | Période : 7 derniers jours"
    c.font = subtitle_font

    ws.merge_cells(f"A3:{last_col_letter}3")
    c = ws["A3"]
    c.value = f"Total : {total_count} événement(s)"
    c.font = summary_font

    ws.row_dimensions[4].height = 6

    header_row = 5
    for col_idx, (header, width) in enumerate(
        zip(_XL_HEADERS, _XL_COL_WIDTHS, strict=True), start=1,
    ):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 28

    row_idx = header_row + 1
    action_col = 4
    result_col = 6

    for log in query.yield_per(500):
        actor = _resolve_actor(log)
        dt = log.created_at
        action_label = ACTION_LABELS_FR.get(log.action_type, log.action_type)
        cat_label = CATEGORY_LABELS_FR.get(log.action_category, log.action_category or "")
        result_label = RESULT_LABELS_FR.get(log.result_status, log.result_status or "")

        values = [
            dt.strftime("%d.%m.%Y") if dt else "",
            dt.strftime("%H:%M") if dt else "",
            actor,
            action_label,
            cat_label,
            result_label,
            mask_ip(log.ip_address),
            parse_device(log.user_agent),
        ]

        is_even = (row_idx - header_row) % 2 == 0
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = cell_align

            if is_even:
                cell.fill = even_fill

            if col_idx == action_col:
                cell.font = data_font_action
            elif col_idx == result_col:
                if log.result_status == "failure":
                    cell.font = result_font_failure
                elif log.result_status == "success":
                    cell.font = result_font_success
                else:
                    cell.font = result_font_other
            elif col_idx in (1, 5, 7):
                cell.font = data_font_secondary
            else:
                cell.font = data_font

        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

    ws.auto_filter.ref = (
        f"A{header_row}:{last_col_letter}{max(row_idx - 1, header_row)}"
    )

    footer_row = row_idx + 1
    ws.merge_cells(f"A{footer_row}:{last_col_letter}{footer_row}")
    c = ws.cell(row=footer_row, column=1)
    c.value = "Les adresses IP sont partiellement masquées pour des raisons de confidentialité."
    c.font = Font(name="Calibri", italic=True, size=9, color=_COLOR_SLATE)

    gen_row = footer_row + 1
    ws.merge_cells(f"A{gen_row}:{last_col_letter}{gen_row}")
    c = ws.cell(row=gen_row, column=1)
    c.value = f"Généré par LIRIE — {now.strftime('%d.%m.%Y %H:%M')}"
    c.font = Font(name="Calibri", italic=True, size=9, color=_COLOR_SLATE)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@security_ns.route("/audit-logs/export")
class AuditLogExport(Resource):
    @jwt_required()
    @role_required(UserRole.company)
    @limiter.limit("5 per hour")
    def get(self):
        """Exporte les audit logs en Excel (.xlsx), CSV ou JSON."""
        company, err, _code = get_company_from_token()
        if err or not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        fmt = request.args.get("format", "xlsx").lower()
        category = request.args.get("category")
        date_from = request.args.get("date_from")
        date_to = request.args.get("date_to")
        raw_limit = min(int(request.args.get("limit", 10000)), 50000)

        query, now, is_institution = _build_export_query(
            company, category, date_from, date_to, raw_limit,
        )

        if not is_institution and fmt == "json":
            fmt = "xlsx"

        from shared.audit_helpers import audit_log
        audit_log(
            "audit_log_exported",
            "security",
            action_details={
                "format": fmt,
                "date_from": date_from,
                "date_to": date_to,
                "category": category,
            },
        )

        ts = now.strftime("%Y%m%d_%H%M%S")
        company_name = getattr(company, "name", "") or "Entreprise"

        if fmt == "xlsx":
            from security.audit_log import AuditLog as AL2
            total_count = AL2.query.filter(AL2.company_id == company.id).count()
            xlsx_bytes = _generate_excel(query, company_name, now, total_count)
            sha256 = hashlib.sha256(xlsx_bytes).hexdigest()
            return Response(
                xlsx_bytes,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="journal_activite_{ts}.xlsx"',
                    "X-Export-SHA256": sha256,
                },
            )

        if fmt == "csv":
            def generate_csv():
                yield _BOM_UTF8 + _CSV_HEADER_LINE
                for batch in query.yield_per(500):
                    yield _format_csv_row(batch)

            return Response(
                generate_csv(),
                mimetype="text/csv; charset=utf-8",
                headers={
                    "Content-Disposition": f'attachment; filename="journal_activite_{ts}.csv"',
                },
            )

        import json as json_mod

        records = []
        for log in query.all():
            records.append(_serialize_audit_log(log))

        body = json_mod.dumps(
            {
                "export_metadata": {
                    "exported_at": now.isoformat(),
                    "exported_by": company_name,
                    "filters": {
                        "date_from": date_from,
                        "date_to": date_to,
                        "category": category,
                    },
                    "total_records": len(records),
                },
                "records": records,
            },
            ensure_ascii=False,
            indent=2,
        )
        sha256 = hashlib.sha256(body.encode("utf-8")).hexdigest()
        return Response(
            body,
            mimetype="application/json; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="journal_activite_{ts}.json"',
                "X-Export-SHA256": sha256,
            },
        )


# ==================== Security Policy ====================

_DEFAULT_POLICY = {
    "require_2fa_roles": [],
    "password_expiry_days": None,
    "max_session_days": 30,
    "enforcement_mode": "warn",
}

security_policy_model = security_ns.model(
    "SecurityPolicy",
    {
        "require_2fa_roles": fields.List(fields.String, description="Roles devant activer 2FA"),
        "password_expiry_days": fields.Integer(description="Jours avant expiration MDP (null=jamais)"),
        "max_session_days": fields.Integer(description="Duree max session en jours"),
        "enforcement_mode": fields.String(description="warn ou enforce"),
    },
)


@security_ns.route("/policy")
class SecurityPolicy(Resource):
    @jwt_required()
    @role_required(UserRole.company)
    def get(self):
        """Retourne la politique de securite de l'entreprise."""
        import json as json_mod

        company, err, _code = get_company_from_token()
        if err or not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        policy = _DEFAULT_POLICY.copy()
        if company.security_policy:
            with suppress(ValueError, TypeError):
                saved = json_mod.loads(company.security_policy)
                policy.update(saved)

        return {"policy": policy}, 200

    @jwt_required()
    @role_required(UserRole.company)
    @security_ns.expect(security_policy_model)
    def put(self):
        """Met a jour la politique de securite de l'entreprise."""
        import json as json_mod

        company, err, _code = get_company_from_token()
        if err or not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        data = request.get_json() or {}

        policy = _DEFAULT_POLICY.copy()
        if company.security_policy:
            with suppress(ValueError, TypeError):
                saved = json_mod.loads(company.security_policy)
                policy.update(saved)

        allowed_keys = {"require_2fa_roles", "password_expiry_days", "max_session_days", "enforcement_mode"}
        for key in allowed_keys:
            if key in data:
                policy[key] = data[key]

        if policy.get("enforcement_mode") not in ("warn", "enforce"):
            policy["enforcement_mode"] = "warn"
        if not isinstance(policy.get("require_2fa_roles"), list):
            policy["require_2fa_roles"] = []

        company.security_policy = json_mod.dumps(policy)
        db.session.commit()

        from shared.audit_helpers import audit_log
        audit_log("security_policy_updated", "security", resource_type="security_policy", resource_id=company.id)

        return {"policy": policy, "message": "Politique de sécurité mise à jour."}, 200


# ==================== Security Alerts (Sprint 3) ====================

alert_item = security_ns.model(
    "SecurityAlert",
    {
        "id": fields.Integer,
        "alert_type": fields.String,
        "message": fields.String,
        "ip_masked": fields.String,
        "device": fields.String,
        "count": fields.Integer,
        "created_at": fields.DateTime,
    },
)


_FAILED_LOGIN_ALERT_THRESHOLD = 3


@security_ns.route("/alerts")
class SecurityAlerts(Resource):
    @jwt_required()
    @role_required(UserRole.company)
    def get(self):
        """Retourne les alertes de securite recentes (30 derniers jours).

        Agrege les tentatives de connexion echouees par fenetre de 1h.
        """
        from shared.security_helpers import mask_ip, parse_device

        company, err, _code = get_company_from_token()
        if err or not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        from security.audit_log import AuditLog

        cutoff = datetime.now(UTC) - timedelta(days=30)
        failed_logins = (
            AuditLog.query.filter(
                AuditLog.company_id == company.id,
                AuditLog.action_type.in_(["login_failed", "user_login_failed"]),
                AuditLog.created_at >= cutoff,
            )
            .order_by(AuditLog.created_at.desc())
            .limit(100)
            .all()
        )

        alerts = []
        if len(failed_logins) >= _FAILED_LOGIN_ALERT_THRESHOLD:
            for log in failed_logins[:10]:
                alerts.append({
                    "id": log.id,
                    "alert_type": "failed_login",
                    "message": "Tentative de connexion échouée",
                    "ip_masked": mask_ip(log.ip_address),
                    "device": parse_device(log.user_agent),
                    "count": 1,
                    "created_at": log.created_at.isoformat() if log.created_at else None,
                })

        return {
            "alerts": alerts,
            "total_failed_logins_30d": len(failed_logins),
        }, 200


@security_ns.route("/alerts/preferences")
class SecurityAlertPreferences(Resource):
    @jwt_required()
    @role_required(UserRole.company)
    def get(self):
        """Retourne les preferences d'alertes email."""
        import json as json_mod

        company, err, _code = get_company_from_token()
        if err or not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        policy = {}
        if company.security_policy:
            with suppress(ValueError, TypeError):
                policy = json_mod.loads(company.security_policy)

        prefs = policy.get("alert_preferences", {
            "failed_logins_burst": True,
            "new_device_login": True,
            "new_country_login": False,
        })
        return {"preferences": prefs}, 200

    @jwt_required()
    @role_required(UserRole.company)
    def put(self):
        """Met a jour les preferences d'alertes email."""
        import json as json_mod

        company, err, _code = get_company_from_token()
        if err or not company:
            return APIErrorHandler.handle_not_found("Company", None, logger)

        data = request.get_json() or {}

        policy = {}
        if company.security_policy:
            with suppress(ValueError, TypeError):
                policy = json_mod.loads(company.security_policy)

        allowed_prefs = {"failed_logins_burst", "new_device_login", "new_country_login"}
        prefs = policy.get("alert_preferences", {})
        for key in allowed_prefs:
            if key in data:
                prefs[key] = bool(data[key])

        policy["alert_preferences"] = prefs
        company.security_policy = json_mod.dumps(policy)
        db.session.commit()

        return {"preferences": prefs, "message": "Préférences mises à jour."}, 200
